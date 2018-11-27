# Coltin Lappin-Lux (Data Analyst)
# lux.coltin@gmail.com
# Boys and Girls Club of Hawaii
# 11/25/18

from __future__ import division
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook

#==================================================================
# Data Structure = List
#==================================================================

def list_insert(sheet):
    myList = []

    col = 'A'
    row = 2
    i = col + str(row)
    while sheet[i].value != None:
        myList.append(sheet[i].value)
        row += 1
        i = col + str(row)

    row_count = len(myList)
    myList = []

    col = 'A'
    row = 1
    i = col + str(row)
    while sheet[i].value != None:
        myList.append(sheet[i].value)
        col = chr(ord(col) + 1)
        i = col + str(row)

    col_count = len(myList)
    myList = []

    mySubList = []

    col = 'A'
    row = 2
    i = col + str(row)
    for k in range(row_count):
        for l in range(col_count):
            #print(sheet[i].value)
            mySubList.append(sheet[i].value)
            col = chr(ord(col) + 1)
            i = col + str(row)
        myList.append(mySubList)
        mySubList = []
        col = 'A'
        row += 1
        i = col + str(row)

    return myList

#==================================================================
# Primary Attribute
#==================================================================

def primary_attribute(data,col):
    #Where to find attribute (Column and Row) Ex. 'B', 2
    unduplicated_list = []

    attribute = ord(col)-65

    for entry in data:
        if unduplicated_list.count(entry[attribute]) == 0:
            unduplicated_list.append(entry[attribute])

    unduplicated_list.sort()

    return unduplicated_list

#==================================================================
# Count Attribute
#==================================================================

def count_attribute(data,primary,col):
    attribute_list = []
    data_count = []
    attribute_count = []
    count = 0
    total = 0

    attribute = ord(col)-65

    for i in range(len(data)):
        if data[i][attribute] == None and attribute_list.count("Unknown") == 0:
            attribute_list.append("Unknown")
        if data[i][attribute] != None and attribute_list.count(data[i][attribute]) == 0:
            attribute_list.append(data[i][attribute])
    attribute_list.sort()

    if attribute_list == primary:
        for l in primary:
            for i in range(len(data)):
                if data[i][1] == l:
                    total += data[i][attribute].count(l)
        for l in primary:
            for i in range(len(data)):
                if data[i][1] == l:
                    count += data[i][attribute].count(l)
            data_count.append([l,[["Membership Count",count],["Membership Percentage",count/total]]])
            count = 0
    else:
        for l in primary:
            for k in attribute_list:
                for i in range(len(data)):
                    if data[i][1] == l:
                        if k == "Unknown" and data[i][attribute] == None:
                            count += 1
                        elif type(data[i][attribute]) != type(None):
                            count += data[i][attribute].count(k)
                attribute_count.append([k,count])
                count = 0
            data_count.append([l,attribute_count])
            attribute_count = []

    return data_count

#==================================================================
# Count Attribute
#==================================================================

def print_to_excel(revisedSheet,data_count,title_index):
    index_counter = title_index
    char_index = 1
    attribute_list = []

    for i in data_count:
        for l in i[1]:
            if attribute_list.count(l[0]) == 0:
                attribute_list.append(l[0])
    
    #Title
    club_i = 'A' + str(index_counter)
    revisedSheet[club_i] = "Sites:"

    for i in range(len(attribute_list)):
        club_i = chr(ord('A') + char_index) + str(index_counter)
        char_index += 1
        revisedSheet[club_i] = attribute_list[i]

    index_counter += 1
    char_index = 1

    for club in data_count:
        club_i = 'A' + str(index_counter)
        revisedSheet[club_i] = club[0]
        for att in club[1]:
            club_i = chr(ord('A') + char_index) + str(index_counter)
            revisedSheet[club_i] = att[1]
            char_index += 1
        index_counter += 1
        char_index = 1

    #total (#)
    club_i = 'A' + str(index_counter)
    revisedSheet[club_i] = "Total (#)"
    index_counter += 1

    #total (%)
    club_i = 'A' + str(index_counter)
    revisedSheet[club_i] = "Total (%)"

    index_counter += 2

    return index_counter

#==================================================================
# Main Function (Initialize Variables and Call Functions)
#==================================================================

def main():
    """
    This function executes when this file is run as a script.
    """

    #==================================================================
    # Open Workbooks (Input and Output Excel Sheets)
    #==================================================================
    #Reading in existing Workbook
    title = raw_input("Input Excel Title: ")
    title = title + '.xlsx'

    wb = load_workbook(filename = title)

    #Creating Revised Workbook
    rwb = Workbook()
    dest_filename = raw_input("Output Excel Title: ")
    dest_filename = dest_filename + '.xlsx'

    #==================================================================
    # While loop in case user wants to summarize multiple excel sheets
    #==================================================================
    #Data Calculation
    continuetill = 'yes'
    while continuetill == 'yes':
        sheetname = raw_input("Input Sheet Name/Number: ")
        sheet = wb[sheetname]
        revisedSheet = rwb.active
        revisedSheet = rwb.create_sheet(title=sheetname)

        #==================================================================
        # Print Title and other document details to new Excel
        #==================================================================
        #Title
        title_index = 1
        title_loc = 'A' + str(title_index)
        revisedSheet[title_loc] = 'Boys & Girls Club of Hawaii'
        #Report Type
        title_index += 1
        title_loc = 'A' + str(title_index)
        revisedSheet[title_loc] = 'Report Type:'
        #Period
        title_index += 1
        title_loc = 'A' + str(title_index)
        revisedSheet[title_loc] = 'Period: '
        #Description
        title_index += 1
        title_loc = 'A' + str(title_index)
        revisedSheet[title_loc] = 'Preparation Date: '

        #==================================================================
        # Storing in Data Structure
        #==================================================================

        data = list_insert(sheet)
        primary_col = raw_input("Primary Attribute Column: ")
        primary_att = primary_attribute(data,primary_col)

        #==================================================================
        # Summarizing Membership
        #==================================================================
        title_index += 2
        data_count = count_attribute(data,primary_att,'B')
        title_index = print_to_excel(revisedSheet,data_count,title_index)

        #==================================================================
        # Summarizing Attributes
        #==================================================================
        att_continuetill = raw_input("Is there another attribute to summarize? (yes/no) ")
        while att_continuetill == 'yes':

            att_col = raw_input("Attribute Column (Letter): ")
            att_row = 2

            data_count = count_attribute(data,primary_att,att_col)
            title_index = print_to_excel(revisedSheet,data_count, title_index)

            att_continuetill = raw_input("Is there another attribute to summarize? (yes/no) ")

        continuetill = raw_input("Is there another sheet? (yes/no) ")

    #Done, Save Revised Workbook
    rwb.save(filename = dest_filename)

if __name__ == "__main__":
    main()