# Coltin Lappin-Lux
# lux.coltin@gmail.com
# 9/18/2019

from __future__ import division
from Tkinter import *
from openpyxl import Workbook
from openpyxl import load_workbook

from Report import *

#==================================================================
# Main Function
#==================================================================

def main():
    def click():
        #==================================================================
        # Load Excel Workbook
        #==================================================================
        #Reading in existing Workbook
        print '\nLoading Excel Workbook ...'
        status.delete(1.0, END)
        status.insert(END, "Loading Excel Workbook ...")
        path = '/Users/coltinlappinlux/Desktop/'
        wbTitle = wbTitleInput.get()
        wbTitle += '.xlsx'
        finalPath = path + wbTitle
        workBook = load_workbook(finalPath)

        #==================================================================
        # Load Excel Sheet & Create Summary Sheet
        #==================================================================
        print '\nLoading Excel Sheet ...'
        status.insert(END, "\nLoading Excel Sheet ...")
        dataSheetName = dataSheetNameInput.get()
        dataSheet = workBook[dataSheetName]
        summarySheetName = dataSheetName + 'Sum'
        summarySheet = workBook.create_sheet(title=summarySheetName)

        #==================================================================
        # Create Report Object
        #==================================================================
        print '\nLoading Report ...'
        status.insert(END, "\nLoading Report ...")
        report = Report()
        report.setDataSheet(dataSheet)
        report.setSummarySheet(summarySheet)

        #==================================================================
        # Create Report Object & Load Report Header
        #==================================================================
        print '\nLoading Report Header ...'
        status.insert(END, "\nLoading Report Header ...")
        companyName = companyInput.get()
        typeName = typeInput.get()
        startDate = startInput.get()
        endDate = endInput.get()
        report.setHeader(companyName, typeName, startDate, endDate)

        #==================================================================
        # Load Report Metadata
        #==================================================================
        print '\nLoading Report Metadata ...'
        status.insert(END, "\nLoading Report Metadata ...")
        primary = primaryInput.get()
        summary = summaryInput.get()
        report.setMetadata(primary, summary)

        #==================================================================
        # Summarize Data
        #==================================================================
        print '\nGenerating Report ...'
        status.insert(END, "\nGenerating Report ...")
        report.generateContent()
        report.printReport(True,True)
        
        workBook.save(finalPath)
        status.insert(END, "\nReport Complete")
    
    def closeWindow():
        window.destroy()
        exit()
    
    def clearValues():
        dataSheetNameInput.delete(0, END)
        typeInput.delete(0, END)
        summaryInput.delete(0, END)
        status.delete(1.0, END)
    
    def clearAllValues():
        wbTitleInput.delete(0, END)
        dataSheetNameInput.delete(0, END)
        companyInput.delete(0, END)
        typeInput.delete(0, END)
        startInput.delete(0, END)
        endInput.delete(0, END)
        primaryInput.delete(0, END)
        summaryInput.delete(0, END)
        status.delete(1.0, END)
    
    window = Tk()
    window.title("Excel Data Summary")
    window.configure(background="white")

    Label(window, text="Excel Title: ", width=19, bg="white", fg="black", font="none 12 bold", anchor=E).grid(row=1, column=0)
    wbTitleInput = Entry(window, width=50, bg="white", fg="black")
    wbTitleInput.grid(row=1,column=1)

    Label(window, text="Excel Sheet: ", width=19, bg="white", fg="black", font="none 12 bold", anchor=E).grid(row=2, column=0)
    dataSheetNameInput = Entry(window, width=50, bg="white", fg="black")
    dataSheetNameInput.grid(row=2,column=1)

    Label(window, text="Company: ", width=19, bg="white", fg="black", font="none 12 bold", anchor=E).grid(row=4, column=0)
    companyInput = Entry(window, width=50, bg="white", fg="black")
    companyInput.grid(row=4,column=1)

    Label(window, text="Report Type: ", width=19, bg="white", fg="black", font="none 12 bold", anchor=E).grid(row=5, column=0)
    typeInput = Entry(window, width=50, bg="white", fg="black")
    typeInput.grid(row=5,column=1)

    Label(window, text="Start Date: ", width=19, bg="white", fg="black", font="none 12 bold", anchor=E).grid(row=6, column=0)
    startInput = Entry(window, width=50, bg="white", fg="black")
    startInput.grid(row=6,column=1)

    Label(window, text="End Date: ", width=19, bg="white", fg="black", font="none 12 bold", anchor=E).grid(row=7, column=0)
    endInput = Entry(window, width=50, bg="white", fg="black")
    endInput.grid(row=7,column=1)

    Label(window, text="Primary Attribute: ", width=19, bg="white", fg="black", font="none 12 bold", anchor=E).grid(row=9, column=0)
    primaryInput = Entry(window, width=50, bg="white", fg="black")
    primaryInput.grid(row=9,column=1)

    Label(window, text="Summary Attribute(s): ", width=19, bg="white", fg="black", font="none 12 bold", anchor=E).grid(row=10, column=0)
    summaryInput = Entry(window, width=50, bg="white", fg="black")
    summaryInput.grid(row=10,column=1)

    status = Text(window, height=7, fg="white", bg="black", font="none 12 bold")
    status.tag_configure("center", justify='center')
    status.grid(row=11, column=0, columnspan=2)

    Button(window, text="Submit", width=13, bg="white", activebackground="green", fg="black", font="none 12 bold", command=click).grid(row=12, column=0, columnspan=2)
    Button(window, text="Clear Part", width=13, bg="white", fg="black", font="none 12 bold", command=clearValues).grid(row=13, column=0, columnspan=2)
    Button(window, text="Clear All", width=13, bg="white", fg="black", font="none 12 bold", command=clearAllValues).grid(row=14, column=0, columnspan=2)
    Button(window, text="Exit", width=13, bg="white", fg="black", font="none 12 bold", command=closeWindow).grid(row=15, column=0, columnspan=2)

    window.mainloop()
    
if __name__ == "__main__":
    main()